"""Views de formação: Cursos, Ações e Inscrições com export XLSX."""
import io
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.core.cache import cache
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from entidades.models import Curso, Acao, Inscricao, ESTADO_PROFISSIONAL_CHOICES, ESTADO_PAGAMENTO_CHOICES
from entidades.forms import InscricaoForm
from .tasks import exportar_inscricoes_async
from django_celery_results.models import TaskResult


def _export_xlsx(nome, cabecalhos, linhas, texto_idx=None):
    wb = Workbook()
    ws = wb.active
    ws.title = nome
    fill = PatternFill("solid", fgColor="1A3C5E")
    fnt  = Font(bold=True, color="FFFFFF")
    for ci, h in enumerate(cabecalhos, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = fill; c.font = fnt; c.alignment = Alignment(horizontal='center')
    for ri, row in enumerate(linhas, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=str(val) if val is not None else '')
            if texto_idx and ci in texto_idx:
                cell.number_format = '@'
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or '')) for c in col) + 4, 50)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{nome}.xlsx"'
    return resp


# ─── AÇÕES ────────────────────────────────────────────────────────────────────

@login_required
@permission_required('entidades.view_acao', raise_exception=True)
def acao_lista(request):
    q = request.GET.get('q', '')
    ano = request.GET.get('ano', '')
    formador = request.GET.get('formador', '')
    qs = Acao.objects.select_related('curso', 'formador').annotate(n_inscricoes=Count('inscricoes'))
    if q:
        qs = qs.filter(Q(referencia__icontains=q) | Q(local__icontains=q))
    if ano:
        qs = qs.filter(ano=ano)
    if formador:
        qs = qs.filter(formador__nome__icontains=formador)

    anos = cache.get_or_set('acao_anos_list', Acao.objects.values_list('ano', flat=True).distinct().order_by('ano'), 3600)

    paginator = Paginator(qs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.GET.get('export') == 'xlsx':
        cabecalhos = ['Refª Ação', 'Curso', 'Local', 'Data Início', 'Data Fim', 'Ano', 'Formador', 'Nº Inscrições']
        linhas = ((
            a.referencia, str(a.curso or ''), a.local,
            str(a.data_inicio or ''), str(a.data_fim or ''),
            a.ano, str(a.formador or ''), a.n_inscricoes
        ) for a in qs.iterator(chunk_size=500))
        return _export_xlsx('acoes', cabecalhos, linhas, texto_idx={1})

    context = {
        'acoes': page_obj, 
        'q': q, 'ano': ano, 'anos': anos, 'formador': formador,
        'page_title': 'Ações', 'total': qs.count(),
        'is_paginated': page_obj.has_other_pages(),
    }
    return render(request, 'formacao/acao_lista.html', context)


@login_required
@permission_required('entidades.view_acao', raise_exception=True)
def acao_detalhe(request, pk):
    acao = get_object_or_404(Acao, pk=pk)
    inscricoes = acao.inscricoes.select_related('formando', 'empresa').order_by('formando__nome')

    if request.GET.get('export') == 'xlsx':
        cabecalhos = ['Formando', 'NIF', 'Empresa', 'Estado Profissional', 'Nº Certificado', 'Estado Pagamento', 'Comercial']
        linhas = ((
            i.formando.nome, i.formando.nif or '',
            str(i.empresa or ''), i.estado_profissional,
            i.numero_certificado, i.estado_pagamento, i.comercial
        ) for i in inscricoes.iterator(chunk_size=500))
        return _export_xlsx(f'acao_{acao.referencia}', cabecalhos, linhas, texto_idx={2})

    context = {'acao': acao, 'inscricoes': inscricoes, 'page_title': acao.referencia}
    return render(request, 'formacao/acao_detalhe.html', context)


# ─── INSCRIÇÕES ───────────────────────────────────────────────────────────────

@login_required
@permission_required('entidades.view_inscricao', raise_exception=True)
def inscricao_lista(request):
    q = request.GET.get('q', '')
    ano = request.GET.get('ano', '')
    estado = request.GET.get('estado', '')
    pagamento = request.GET.get('pagamento', '')

    qs = Inscricao.objects.select_related('formando', 'acao', 'acao__curso', 'acao__formador', 'empresa')
    if q:
        qs = qs.filter(Q(formando__nome__icontains=q) | Q(formando__nif__icontains=q)
                       | Q(acao__referencia__icontains=q))
    if ano:
        qs = qs.filter(acao__ano=ano)
    if estado:
        qs = qs.filter(estado_profissional=estado)
    if pagamento:
        qs = qs.filter(estado_pagamento=pagamento)

    anos = cache.get_or_set('acao_anos_list', Acao.objects.values_list('ano', flat=True).distinct().order_by('ano'), 3600)

    paginator = Paginator(qs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.GET.get('export') == 'xlsx':
        if qs.count() > 5000:
            messages.warning(request, "Volume de dados elevado. A usar exportação inteligente (Async).")
            filtros = {'q': q, 'ano': ano, 'estado': estado, 'pagamento': pagamento}
            exportar_inscricoes_async.delay(filtros, user_id=request.user.id)
            request.session['export_active'] = True
            messages.success(request, "Exportação Inteligente iniciada! O ficheiro estará disponível em breve.")
            return redirect('formacao:inscricao_lista')
        cabecalhos = ['Formando', 'NIF', 'Empresa', 'Refª Ação', 'Curso', 'Local',
                      'Ano', 'Estado Prof.', 'Nº Cert.', 'Estado Pag.', 'Formador', 'Comercial']
        linhas = ((
            i.formando.nome, i.formando.nif or '',
            str(i.empresa or ''), i.acao.referencia,
            str(i.acao.curso or ''), i.acao.local, i.acao.ano,
            i.estado_profissional, i.numero_certificado,
            i.estado_pagamento, str(i.acao.formador or ''), i.comercial
        ) for i in qs.iterator(chunk_size=500))
        return _export_xlsx('inscricoes', cabecalhos, linhas, texto_idx={2, 4})

    if request.GET.get('export') == 'async':
        filtros = {
            'q': q, 'ano': ano, 
            'estado': estado, 'pagamento': pagamento
        }
        exportar_inscricoes_async.delay(filtros, user_id=request.user.id)
        request.session['export_active'] = True
        messages.success(request, "Exportação Inteligente iniciada! O ficheiro estará disponível no seu histórico de downloads em breve.")
        return redirect('formacao:inscricao_lista')

    context = {
        'inscricoes': page_obj, 'q': q, 'ano': ano, 'anos': anos,
        'estado': estado, 'pagamento': pagamento,
        'page_title': 'Inscrições', 'total': qs.count(),
        'ESTADOS': [e[0] for e in ESTADO_PROFISSIONAL_CHOICES],
        'PAGAMENTOS': [e[0] for e in ESTADO_PAGAMENTO_CHOICES],
        'is_paginated': page_obj.has_other_pages(),
    }
    return render(request, 'formacao/inscricao_lista.html', context)


@login_required
@permission_required('entidades.add_inscricao', raise_exception=True)
def inscricao_criar(request):
    form = InscricaoForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, 'Inscrição registada.')
        return redirect('formacao:inscricao_lista')
    return render(request, 'entidades/form.html', {'form': form, 'page_title': 'Nova Inscrição', 'entidade': 'Inscrição'})


@login_required
@permission_required('entidades.change_inscricao', raise_exception=True)
def inscricao_editar(request, pk):
    inscricao = get_object_or_404(Inscricao, pk=pk)
    form = InscricaoForm(request.POST or None, instance=inscricao)
    if form.is_valid():
        form.save()
        messages.success(request, 'Inscrição atualizada.')
        return redirect('formacao:inscricao_lista')
    return render(request, 'entidades/form.html', {'form': form, 'page_title': 'Editar Inscrição', 'entidade': 'Inscrição'})


@login_required
@permission_required('entidades.view_inscricao', raise_exception=True)
def check_export_status(request):
    """Retorna o estado da última exportação do utilizador via HTMX."""
    tarefa = TaskResult.objects.filter(
        task_name='formacao.tasks.exportar_inscricoes_async',
        task_kwargs__contains=f'"user_id": {request.user.id}'
    ).order_by('-date_created').first()
    
    if not tarefa:
        request.session.pop('export_active', None)
        return HttpResponse("")
    
    if tarefa.status == 'SUCCESS':
        url_download = tarefa.result.strip('"')
        request.session.pop('export_active', None)
        response = HttpResponse(
            f'<a href="{url_download}" class="btn btn-success btn-sm fade-in"><i class="bi bi-download"></i> Download Pronto</a>'
        )
        response.status_code = 286
        return response
    
    if tarefa.status == 'FAILURE':
        request.session.pop('export_active', None)
        response = HttpResponse('<span class="badge bg-danger">Falha na Exportação</span>')
        response.status_code = 286
        return response

    return HttpResponse('<span class="badge bg-info anim-pulse"><i class="bi bi-gear-fill"></i> A processar XLSX...</span>')
