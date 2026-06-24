import os
import uuid
from datetime import datetime
from celery import shared_task
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from entidades.models import Inscricao

@shared_task
def exportar_inscricoes_async(filtros, user_id=None):
    """
    Gera ficheiro XLSX de inscrições em background.
    'filtros' é um dict com: q, ano, estado, pagamento
    'user_id' é o ID do utilizador que solicitou (para filtragem segura)
    """
    qs = Inscricao.objects.select_related('formando', 'acao', 'acao__curso', 'acao__formador', 'empresa').all()
    
    # Aplicar os mesmos filtros da View
    q = filtros.get('q')
    if q:
        from django.db.models import Q
        qs = qs.filter(Q(formando__nome__icontains=q) | Q(formando__nif__icontains=q) | Q(acao__referencia__icontains=q))
    
    ano = filtros.get('ano')
    if ano: qs = qs.filter(acao__ano=ano)
    
    estado = filtros.get('estado')
    if estado: qs = qs.filter(estado_profissional=estado)
    
    pagamento = filtros.get('pagamento')
    if pagamento: qs = qs.filter(estado_pagamento=pagamento)

    # 1. Preparar Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Inscricoes"
    
    fill = PatternFill("solid", fgColor="1A3C5E")
    fnt  = Font(bold=True, color="FFFFFF")
    
    cabecalhos = [
        'Formando', 'NIF', 'Empresa', 'Refª Ação', 'Curso', 'Local',
        'Ano', 'Estado Prof.', 'Nº Cert.', 'Estado Pag.', 'Formador', 'Comercial'
    ]
    
    for ci, h in enumerate(cabecalhos, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = fill; c.font = fnt; c.alignment = Alignment(horizontal='center')

    # 2. Iterar QuerySet (Gerador para poupar memória)
    for ri, i in enumerate(qs.iterator(), 2):
        row = [
            i.formando.nome, i.formando.nif or '',
            str(i.empresa or ''), i.acao.referencia,
            str(i.acao.curso or ''), i.acao.local, i.acao.ano,
            i.estado_profissional, i.numero_certificado,
            i.estado_pagamento, str(i.acao.formador or ''), i.comercial
        ]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=str(val) if val is not None else '')
            if ci in [2, 4]: # NIF e Ref Ação como texto
                cell.number_format = '@'

    # 3. Ajustar largura automaticamente (limitado a 50 para performance)
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # 4. Salvar ficheiro
    filename = f"inscricoes_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}.xlsx"
    export_path = os.path.join(settings.MEDIA_ROOT, 'exports')
    os.makedirs(export_path, exist_ok=True)
    
    full_path = os.path.join(export_path, filename)
    wb.save(full_path)
    
    # Retorna o URL relativo para o download
    return f"{settings.MEDIA_URL}exports/{filename}"
