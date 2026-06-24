"""
Views de entidades: Empresa, Formador, Formando (CRUD + Export XLSX).
"""
import io
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Empresa, Formador, Formando, Inscricao
from .forms import EmpresaForm, FormadorForm, FormandoForm


# ─── Utilitário de Export XLSX ────────────────────────────────────────────────

COLUNAS_TEXTO = {'nif', 'numero_identificacao', 'codigo_postal', 'telefone1', 'telefone2'}

def _export_xlsx(nome_ficheiro, cabecalhos, linhas, colunas_texto_idx=None):
    """Gera um ficheiro Excel seguro: NIFs e códigos sempre como texto."""
    wb = Workbook()
    ws = wb.active
    ws.title = nome_ficheiro

    # Estilo do cabeçalho
    header_fill = PatternFill("solid", fgColor="1A3C5E")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, cab in enumerate(cabecalhos, 1):
        cell = ws.cell(row=1, column=col_idx, value=cab)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Dados
    for row_idx, linha in enumerate(linhas, 2):
        for col_idx, valor in enumerate(linha, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(valor) if valor is not None else '')
            # Forçar formato Texto para colunas críticas (NIFs, códigos)
            if colunas_texto_idx and col_idx in colunas_texto_idx:
                cell.number_format = '@'

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nome_ficheiro}.xlsx"'
    return response


# ─── EMPRESAS ─────────────────────────────────────────────────────────────────

@login_required
@permission_required('entidades.view_empresa', raise_exception=True)
def empresa_lista(request):
    q = request.GET.get('q', '')
    localidade = request.GET.get('localidade', '')
    qs = Empresa.objects.annotate(total_formandos=Count('formandos'))
    if q:
        qs = qs.filter(Q(nome__icontains=q) | Q(nif__icontains=q))
    if localidade:
        qs = qs.filter(localidade__icontains=localidade)

    paginator = Paginator(qs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.GET.get('export') == 'xlsx':
        cabecalhos = ['Nome', 'NIF', 'Morada', 'Código Postal', 'Localidade']
        linhas = ((e.nome, e.nif, e.morada, e.codigo_postal, e.localidade) for e in qs.iterator(chunk_size=500))
        return _export_xlsx('empresas', cabecalhos, linhas, colunas_texto_idx={2})

    context = {
        'empresas': page_obj,
        'q': q,
        'localidade': localidade,
        'page_title': 'Empresas',
        'total': qs.count(),
        'is_paginated': page_obj.has_other_pages(),
    }
    return render(request, 'entidades/empresa_lista.html', context)


@login_required
@permission_required('entidades.view_empresa', raise_exception=True)
def empresa_detalhe(request, pk):
    empresa = get_object_or_404(Empresa, pk=pk)
    formandos_qs = empresa.formandos.select_related('empresa')
    paginator = Paginator(formandos_qs, 50)
    page_number = request.GET.get('page')
    formandos = paginator.get_page(page_number)
    context = {
        'empresa': empresa, 'formandos': formandos,
        'is_paginated': formandos.has_other_pages(),
        'page_title': empresa.nome,
    }
    return render(request, 'entidades/empresa_detalhe.html', context)


@login_required
@permission_required('entidades.add_empresa', raise_exception=True)
@transaction.atomic
def empresa_criar(request):
    form = EmpresaForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, 'Empresa criada com sucesso.')
        return redirect('entidades:empresa_lista')
    return render(request, 'entidades/form.html', {'form': form, 'page_title': 'Nova Empresa', 'entidade': 'Empresa'})


@login_required
@permission_required('entidades.change_empresa', raise_exception=True)
@transaction.atomic
def empresa_editar(request, pk):
    empresa = get_object_or_404(Empresa, pk=pk)
    form = EmpresaForm(request.POST or None, instance=empresa)
    if form.is_valid():
        form.save()
        messages.success(request, 'Empresa atualizada.')
        return redirect('entidades:empresa_detalhe', pk=pk)
    return render(request, 'entidades/form.html', {'form': form, 'page_title': f'Editar {empresa.nome}', 'entidade': 'Empresa'})


@login_required
@permission_required('entidades.delete_empresa', raise_exception=True)
def empresa_eliminar(request, pk):
    empresa = get_object_or_404(Empresa, pk=pk)
    empresa.delete()
    messages.success(request, f'Empresa "{empresa.nome}" eliminada.')
    return redirect('entidades:empresa_lista')


# ─── FORMADORES ───────────────────────────────────────────────────────────────

@login_required
@permission_required('entidades.view_formador', raise_exception=True)
def formador_lista(request):
    q = request.GET.get('q', '')
    qs = Formador.objects.annotate(total_acoes=Count('acoes'))
    if q:
        qs = qs.filter(Q(nome__icontains=q) | Q(email1__icontains=q))

    paginator = Paginator(qs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.GET.get('export') == 'xlsx':
        cabecalhos = ['ID', 'Nome', 'Código', 'Telefone', 'Email', 'Morada']
        linhas = ((f.pk, f.nome, f.codigo, f.telefone1, f.email1, f.morada) for f in qs.iterator(chunk_size=500))
        return _export_xlsx('formadores', cabecalhos, linhas)

    context = {
        'formadores': page_obj, 
        'q': q, 
        'page_title': 'Formadores', 
        'total': qs.count(),
        'is_paginated': page_obj.has_other_pages(),
    }
    return render(request, 'entidades/formador_lista.html', context)


@login_required
@permission_required('entidades.view_formador', raise_exception=True)
def formador_detalhe(request, pk):
    formador = get_object_or_404(Formador, pk=pk)
    acoes = formador.acoes.select_related('curso').order_by('-ano')[:30]
    context = {'formador': formador, 'acoes': acoes, 'page_title': formador.nome}
    return render(request, 'entidades/formador_detalhe.html', context)


@login_required
@permission_required('entidades.add_formador', raise_exception=True)
@transaction.atomic
def formador_criar(request):
    form = FormadorForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, 'Formador criado com sucesso.')
        return redirect('entidades:formador_lista')
    return render(request, 'entidades/form.html', {'form': form, 'page_title': 'Novo Formador', 'entidade': 'Formador'})


@login_required
@permission_required('entidades.change_formador', raise_exception=True)
@transaction.atomic
def formador_editar(request, pk):
    formador = get_object_or_404(Formador, pk=pk)
    form = FormadorForm(request.POST or None, instance=formador)
    if form.is_valid():
        form.save()
        messages.success(request, 'Formador atualizado.')
        return redirect('entidades:formador_detalhe', pk=pk)
    return render(request, 'entidades/form.html', {'form': form, 'page_title': f'Editar {formador.nome}', 'entidade': 'Formador'})


@login_required
@permission_required('entidades.delete_formador', raise_exception=True)
def formador_eliminar(request, pk):
    formador = get_object_or_404(Formador, pk=pk)
    formador.delete()
    messages.success(request, f'Formador "{formador.nome}" eliminado.')
    return redirect('entidades:formador_lista')


# ─── FORMANDOS ────────────────────────────────────────────────────────────────

@login_required
@permission_required('entidades.view_formando', raise_exception=True)
def formando_lista(request):
    q = request.GET.get('q', '')
    empresa_id = request.GET.get('empresa', '')
    
    # Subquery para pegar o estado mais recente
    from django.db.models import OuterRef, Subquery
    ultimo_estado = Inscricao.objects.filter(formando=OuterRef('pk')).order_by('-id').values('estado_profissional')[:1]

    qs = Formando.objects.select_related('empresa').annotate(
        total_inscricoes=Count('inscricoes'),
        ultimo_estado=Subquery(ultimo_estado)
    )
    
    if q:
        qs = qs.filter(Q(nome__icontains=q) | Q(nif__icontains=q))
    if empresa_id:
        qs = qs.filter(empresa_id=empresa_id)

    paginator = Paginator(qs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.GET.get('export') == 'xlsx':
        cabecalhos = ['Nome', 'NIF', 'Empresa', 'Email', 'Telefone', 'Localidade']
        linhas = ((f.nome, f.nif or '', str(f.empresa or ''), f.email, f.telefone, f.localidade) for f in qs.iterator(chunk_size=500))
        return _export_xlsx('formandos', cabecalhos, linhas, colunas_texto_idx={2})

    context = {
        'formandos': page_obj, 
        'q': q, 
        'page_title': 'Formandos', 
        'total': qs.count(),
        'is_paginated': page_obj.has_other_pages(),
    }
    return render(request, 'entidades/formando_lista.html', context)


@login_required
@permission_required('entidades.view_formando', raise_exception=True)
def formando_detalhe(request, pk):
    formando = get_object_or_404(Formando, pk=pk)
    # ML Recomendação
    recomendacao = formando.recomendacoes_ml.select_related('curso_sugerido').first()
    
    inscricoes_qs = (
        formando.inscricoes
        .select_related('acao', 'acao__curso', 'empresa', 'previsao_churn')
        .order_by('-acao__ano')
    )
    paginator = Paginator(inscricoes_qs, 50)
    page_number = request.GET.get('page')
    inscricoes = paginator.get_page(page_number)
    
    context = {
        'formando': formando, 
        'inscricoes': inscricoes, 
        'recomendacao': recomendacao,
        'is_paginated': inscricoes.has_other_pages(),
        'page_title': formando.nome
    }
    return render(request, 'entidades/formando_detalhe.html', context)


@login_required
@permission_required('entidades.add_formando', raise_exception=True)
@transaction.atomic
def formando_criar(request):
    form = FormandoForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, 'Formando criado com sucesso.')
        return redirect('entidades:formando_lista')
    return render(request, 'entidades/form.html', {'form': form, 'page_title': 'Novo Formando', 'entidade': 'Formando'})


@login_required
@permission_required('entidades.change_formando', raise_exception=True)
@transaction.atomic
def formando_editar(request, pk):
    formando = get_object_or_404(Formando, pk=pk)
    form = FormandoForm(request.POST or None, instance=formando)
    if form.is_valid():
        form.save()
        messages.success(request, 'Formando atualizado.')
        return redirect('entidades:formando_detalhe', pk=pk)
    return render(request, 'entidades/form.html', {'form': form, 'page_title': f'Editar {formando.nome}', 'entidade': 'Formando'})


@login_required
@permission_required('entidades.delete_formando', raise_exception=True)
def formando_eliminar(request, pk):
    formando = get_object_or_404(Formando, pk=pk)
    formando.delete()
    messages.success(request, f'Formando "{formando.nome}" eliminado.')
    return redirect('entidades:formando_lista')
